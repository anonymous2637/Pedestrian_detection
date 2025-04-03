import cv2
import time
import sqlite3
import pandas as pd
from datetime import datetime
from ultralytics import YOLO
from openpyxl import load_workbook

# RTSP Camera URL
ip_camera_url = "rtsp://admin:admin123@192.168.1.213/cam/realmonitor?channel=1&subtype=0"

# Load YOLO model
model = YOLO("yolov5mu.pt")  
confidence_threshold = 0.3  

# SQLite Database Connection
conn = sqlite3.connect("detections.db", check_same_thread=False)
cursor = conn.cursor()
cursor.execute("""
    CREATE TABLE IF NOT EXISTS person_detections (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        time TEXT,
        person_count INTEGER
    )
""")
conn.commit()

# Excel file setup
excel_file = "detections.xlsx"
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    df = pd.DataFrame(columns=["Date", "Time", "Person Count"])
    df.to_excel(excel_file, index=False, sheet_name="Detections")
    workbook = load_workbook(excel_file)
    sheet = workbook.active

# Function to save detections
def save_detection(person_count):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time_12hr = now.strftime("%I:%M:%S %p")

    cursor.execute("INSERT INTO person_detections (date, time, person_count) VALUES (?, ?, ?)",
                   (date, time_12hr, person_count))
    conn.commit()

    # Update Excel file
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([date, time_12hr, person_count])
        workbook.save(excel_file)
        workbook.close()
    except Exception as e:
        print(f"Error updating Excel file: {e}")

# Function to connect to the camera
def connect_camera(url):
    cap = cv2.VideoCapture(url, cv2.CAP_FFMPEG)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    return cap

# Open camera connection
cap = connect_camera(ip_camera_url)
last_detection_time = 0  

detection_enabled = True  # Enable detection by default

# Function to capture video frames
def process_frames():
    global detection_enabled, last_detection_time
    global cap

    cv2.namedWindow("IP Camera Feed", cv2.WND_PROP_FULLSCREEN)
    cv2.setWindowProperty("IP Camera Feed", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

    while True:
        for _ in range(5):  # Drop outdated frames
            cap.grab()

        success, frame = cap.read()
        if not success:
            print("Lost connection. Reconnecting...")
            cap.release()
            time.sleep(1)
            cap = connect_camera(ip_camera_url)
            continue

        if detection_enabled:
            results = model(frame, conf=confidence_threshold)
            for result in results:
                people = [box for box in result.boxes if int(box.cls[0]) == 0 and box.conf[0] > confidence_threshold]
                person_count = len(people)
                for box in people:
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(frame, "Person", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

                if person_count > 0 and time.time() - last_detection_time > 2:
                    save_detection(person_count)
                    last_detection_time = time.time()

        # Display the live feed in full screen
        cv2.imshow("IP Camera Feed", frame)

        # Press 'q' to exit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    process_frames()
