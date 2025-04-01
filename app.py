from flask import Flask, render_template, Response, request, jsonify, redirect, url_for
import cv2
import time
import sqlite3
import pandas as pd
from datetime import datetime
from ultralytics import YOLO
from openpyxl import load_workbook

app = Flask(__name__)

# RTSP Camera URL
ip_camera_url = "rtsp://admin:admin123@192.168.1.213/cam/realmonitor?channel=1&subtype=0"

# Load YOLO model
model = YOLO("yolov5mu.pt")  
confidence_threshold = 0.3  

# SQLite Database Connection
conn = sqlite3.connect("detections.db", check_same_thread=False)
cursor = conn.cursor()
cursor.execute("""
    CREATE TABLE IF NOT EXISTS person_detections (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        time TEXT,
        person_count INTEGER
    )
""")
conn.commit()

# Excel file setup
excel_file = "detections.xlsx"
try:
    with open(excel_file, "rb"):
        existing_file = True
except FileNotFoundError:
    existing_file = False

if not existing_file:
    df = pd.DataFrame(columns=["Date", "Time", "Person Count"])
    df.to_excel(excel_file, index=False, sheet_name="Detections")

# Variable to control detection state
detection_enabled = False

# Function to save detections
def save_detection(person_count):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time_12hr = now.strftime("%I:%M:%S %p")

    cursor.execute("INSERT INTO person_detections (date, time, person_count) VALUES (?, ?, ?)",
                   (date, time_12hr, person_count))
    conn.commit()

    # Update Excel file
    df = pd.DataFrame([[date, time_12hr, person_count]], columns=["Date", "Time", "Person Count"])
    
    with pd.ExcelWriter(excel_file, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        df.to_excel(writer, index=False, header=False, sheet_name="Detections", startrow=sheet.max_row)

# Function to connect to the camera
def connect_camera(url):
    cap = cv2.VideoCapture(url, cv2.CAP_FFMPEG)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    return cap, width, height

# Open camera connection
cap, width, height = connect_camera(ip_camera_url)
last_detection_time = 0  

# Function to capture video frames
def generate_frames():
    global detection_enabled, last_detection_time
    global cap, width, height

    while True:
        for _ in range(5):  # Drop outdated frames
            cap.grab()

        success, frame = cap.read()
        if not success:
            print("Lost connection. Reconnecting...")
            cap.release()
            time.sleep(1)
            cap, width, height = connect_camera(ip_camera_url)
            continue

        frame = cv2.resize(frame, (width, height))
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        if detection_enabled:
            results = model(cv2.cvtColor(frame, cv2.COLOR_RGB2BGR), conf=confidence_threshold)
            for result in results:
                people = [box for box in result.boxes if int(box.cls[0]) == 0 and box.conf[0] > confidence_threshold]
                person_count = len(people)
                for box in people:
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(frame, "Person", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

                if person_count > 0 and time.time() - last_detection_time > 2:
                    save_detection(person_count)
                    last_detection_time = time.time()

        _, buffer = cv2.imencode('.jpg', cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
        frame = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/detection')
def detection():
    return render_template('detection.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/toggle_detection', methods=['POST'])
def toggle_detection():
    global detection_enabled
    detection_enabled = not detection_enabled
    return jsonify({"detection_status": detection_enabled})

if __name__ == "__main__":
    app.run(debug=True)